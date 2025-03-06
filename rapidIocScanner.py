import os
import threading
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from config import VIRUS_TOTAL_KEYS, ABUSE_IP_KEYS
from parser import args

from colorama import Fore, Style

from VirusTotal.virus_total import VirusTotal
from AbuseIPDb.abuse_ip_db import AbuseIPDb
from utils import check_ioc_type

FILE = args.file
COLUMN = args.column if args.column else "A"
SHEET = args.sheet if args.sheet else 0

def fetch_abuse_ip(key_index: int, value: str, results, new_columns):
    """Fetch data from AbuseIPDb."""
    abuse_ioc = AbuseIPDb(ABUSE_IP_KEYS[key_index], value)
    abuse_ioc.fetch_data()

    score = abuse_ioc.get_abuse_conf_score()
    link = abuse_ioc.get_link()

    results.update({"AI Score": score, "AI Link": link})
    
    if args.append:
        new_columns["AI Scores"].append(score)

def fetch_virus_total(key_index: int, value, ioc_type, results, new_columns):
    """Fetch data from VirusTotal."""
    virus_ioc = VirusTotal(VIRUS_TOTAL_KEYS[key_index], value, ioc_type)
    virus_ioc.fetch_data()

    score = virus_ioc.get_community_score()
    link = virus_ioc.get_link()
    flag = "Clean" if score == 0 else "Unrated" if score == 404 else "Malicious"

    results.update({"VT Score": score, "VT Link": link, "Flag": flag})

    if args.append:
        new_columns["VT Scores"].append(score)
        new_columns["Flag"].append(flag)

def process_iocs(column_values):
    """Process all IOCs with threading and store results, with a progress bar."""
    results_list = []
    new_columns = {"AI Scores": [], "VT Scores": [], "Flag": []}
    KEYS_POINTER = 0

    with tqdm(total=len(column_values), desc="Processing IOCs", unit="ioc") as pbar:
        for value in column_values:
            threads = []
            ioc_type = check_ioc_type(value)
            results = {"IOC": value, "IOC Type": ioc_type, "Flag": "N/A", "AI Score": "N/A", "VT Score": "N/A", "AI Link": "N/A", "VT Link": "N/A"}

            if KEYS_POINTER == len(VIRUS_TOTAL_KEYS) - 1:
                KEYS_POINTER = 0
            else:
                KEYS_POINTER += 1
            
            if ioc_type == "IP":
                t1 = threading.Thread(target=fetch_abuse_ip, args=(KEYS_POINTER, value, results, new_columns))
                threads.append(t1)
                t1.start()

            t2 = threading.Thread(target=fetch_virus_total, args=(KEYS_POINTER, value, ioc_type, results, new_columns))
            threads.append(t2)
            t2.start()

            for t in threads:
                t.join()

            results_list.append(results)
            pbar.update(1)

    return results_list, new_columns

def write_results(results_list, new_columns):
    """Write results to the Excel file."""
    if args.append and os.path.exists(FILE):
        wb = load_workbook(FILE)
        sheet = wb.worksheets[SHEET]

        max_column = sheet.max_column + 1
        column_letters = [chr(64 + max_column + i) for i in range(len(new_columns))]

        for i, (header, values) in enumerate(new_columns.items()):
            column_letter = column_letters[i]
            sheet[f"{column_letter}1"] = header

            for row_index, value in enumerate(values):
                sheet[f"{column_letter}{row_index + 2}"] = value

        wb.save(FILE)
        print(f"✅ Data appended to {FILE}, Sheet {SHEET}.")
    else:
        output_folder = "output"
        os.makedirs(output_folder, exist_ok=True)

        file_path = os.path.join(output_folder, "ioc_results.xlsx")

        df = pd.DataFrame(results_list)
        df.to_excel(file_path, index=False)
        print("✅ New file created: ioc_results.xlsx")

if __name__ == '__main__':
    banner = """
______            _     _ _____      _____   _____                                 
| ___ \\          (_)   | |_   _|    /  __ \\ /  ___|                                
| |_/ /__ _ _ __  _  __| | | |  ___ | /  \\/ \\ `--.  ___ __ _ _ __  _ __   ___ _ __ 
|    // _` | '_ \\| |/ _` | | | / _ \\| |      `--. \\/ __/ _` | '_ \\| '_ \\ / _ \\ '__|
| |\\ \\ (_| | |_) | | (_| |_| || (_) | \\__/\\ /\\__/ / (_| (_| | | | | | | |  __/ |   
\\_| \\_\\__,_| .__/|_|\\__,_|\\___/\\___/ \\____/ \\____/ \\___\\__,_|_| |_|_| |_|\\___|_|   
           | |                                                                     
           |_|                                                                                                                     
"""
    print(Fore.GREEN + banner + Style.RESET_ALL)
    excel_iocs = pd.read_excel(FILE, sheet_name=SHEET, usecols=COLUMN)
    column_values = excel_iocs.iloc[:, 0].dropna().tolist()
    results_list, new_columns = process_iocs(column_values)
    write_results(results_list, new_columns)
